import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


def export_audit_questions():

    # Load control library
    df = pd.read_excel("questions.xlsx")

    # Ensure required columns exist
    if "Framework" not in df.columns:
        df["Framework"] = ""

    if "Control ID" not in df.columns:
        df["Control ID"] = ""

    # Remove old numbering
    if "S.No" in df.columns:
        df = df.drop(columns=["S.No"])

    # Sort controls by domain
    df = df.sort_values(by="Domain")

    df.reset_index(drop=True, inplace=True)

    # Generate numbering
    df.insert(0, "S.No", range(1, len(df) + 1))

    # Reset audit columns
    if "Status" in df.columns:
        df = df.drop(columns=["Status"])

    df["Status"] = "Not Tested"
    df["Evidence"] = ""
    df["Comments"] = ""

    # Export checklist
    df.to_excel("audit_checklist.xlsx", index=False)

    # Open workbook
    wb = load_workbook("audit_checklist.xlsx")
    ws = wb.active

    # Create dropdown validation
    dv = DataValidation(type="list",
                        formula1='"Pass,Fail,Not Tested"',
                        allow_blank=False)

    ws.add_data_validation(dv)

    # Status column = F
    for row in range(2, len(df) + 2):
        dv.add(ws[f"F{row}"])

    # Define colors
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Conditional formatting
    ws.conditional_formatting.add(
        f"F2:F{len(df)+1}",
        FormulaRule(formula=['F2="Pass"'], fill=green)
    )

    ws.conditional_formatting.add(
        f"F2:F{len(df)+1}",
        FormulaRule(formula=['F2="Fail"'], fill=red)
    )

    ws.conditional_formatting.add(
        f"F2:F{len(df)+1}",
        FormulaRule(formula=['F2="Not Tested"'], fill=yellow)
    )

    wb.save("audit_checklist.xlsx")

    print("Audit checklist created successfully!")


if __name__ == "__main__":
    export_audit_questions()